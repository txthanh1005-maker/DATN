
"""
[Semantic Metadata]
- Implements: Thư viện lõi chứa các hàm cấu trúc, phương trình trào lưu công suất, và các hàm hỗ trợ thiết lập hệ thống lưới điện.
- Related_to_Paper: Emergency_Sharing_Plan, Literature reivew knowleadge
"""

#! Khai báo thư viện 
from fileinput import filename
from pathlib import Path

import math 
from math import sqrt 
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Color
from pyomo.environ import value
#! MUSIC
# import pygame
import time

def play_victory_music():
    # pygame.mixer.init()
    # Bạn hãy để file nhạc cùng thư mục với file code nhé
    # pygame.mixer.music.load("victory_song.mp3") 
    # pygame.mixer.music.play()
    time.sleep(1)
    # pygame.mixer.music.stop()
    # while pygame.mixer.music.get_busy():
    #     time.sleep(1)
    return 

import pyomo.environ as pe
from pyomo.environ import SolverStatus, TerminationCondition
from pyomo.environ import SolverFactory
from pyomo.util.infeasible import log_infeasible_constraints
import logging
oo = 999999999
scale_power = 10000 #! *10000 <=> 10kW
S_base = 1000000 #*(1MW)
U_base = 4800  #*(22kV)
Z_base = U_base*U_base/S_base #*(484 Ohm)
I_base = S_base/U_base 



#! KHAI BÁO BASEDATA CỦA MODEL FOR EACH MG

def Create_Model_basedata(model, number_of_node, number_of_Configuration):
#* hàm để khai báo số node, số pha, số configuration, type of load
  model.index_of_node=pe.RangeSet(0,number_of_node) #n
  model.index_of_Configuration = pe.RangeSet(1,number_of_Configuration) #c
  H = getattr(model, 'H', 24)
  model.time = pe.RangeSet(0, H - 1)
  return model
def Create_Configuration(model, filename):
#*hàm này dùng để tạo ra Configuration với mỗi file đọc vào
  df = pd.read_excel(filename, engine='openpyxl') 
  #*tạo dictionary từ số liệu lấy từ file excel
  #print(df.head())
  
  Line_config= {}

  for n in model.index_of_Configuration:
    r = float(df.loc[n-1, 'r1(ohm/kft)'])
    x = float(df.loc[n-1, 'x1(ohm/kft)'])
    c = float(df.loc[n-1, 'c1 (nF/kft)'])
    Line_config[n]= [r,x,c] 
    
#   print(Line_config)
  return Line_config
def Create_Z_line(model, filename, line_config): 
#* tạo ra Z_line với mỗi file đọc vào
  df = pd.read_excel(filename, engine='openpyxl')
  Line_dict ={}
  R_Line_dict ={}
  X_Line_dict ={}
  C_Line_dict ={}
  #print(df.columns) 
  for idx, value in df['node_from'].items():
      from_value = int(value)
      to_value = int(df.at[idx, 'node_to'])
      config = int(df.at[idx, 'config'])
      length = int(df.at[idx, 'length'])
      length = length/1000 #!? đã chỉnh scale RX về chuẩn 
      if from_value > to_value:
          from_value, to_value = to_value, from_value
      """
      vì đã đổi giá trì ở đây để hạ 50% dung lượng bộ nhớ nên mỗi lần truy xuất phải
      đảm bảo from_value< to_value nếu không sẽ sinh lỗi
      """
      
      Line_dict[(from_value, to_value)] = 0
      R_Line_dict[(to_value)] = line_config[config][0]*length/Z_base
      X_Line_dict[(to_value)] = line_config[config][1]*length/Z_base
      C_Line_dict[(to_value)] = line_config[config][2]*length/Z_base
  R_Line_dict[0] = 0
  X_Line_dict[0] = 0
  C_Line_dict[0] = 0
  R_Line_dict[1] = 0
  X_Line_dict[1] = 0
  C_Line_dict[1] = 0
  model.node_links = pe.Set(dimen=2, initialize=list(Line_dict.keys()))
  #*model này dùng để biểu diễn tất cả các đường có thể đi của lưới

  '''
  FINAL LINE_DATA
  '''
   #* NOTE: các Z line ij ở đây đang được quy về node sau của nó là Z_line[j]
  model.R_line= pe.Param(model.index_of_node, initialize=R_Line_dict)
  model.X_line= pe.Param(model.index_of_node, initialize=X_Line_dict)
  model.C_line= pe.Param(model.index_of_node, initialize=C_Line_dict)
  return model
def Create_Load_data(model, number_of_node,MG_index):
#* tạo ra lOAD từng node với mỗi file đọc vào
  model.p_node = pe.Param(model.index_of_node, model.time,initialize = 0, within = pe.Reals, mutable = True )  
  model.q_node = pe.Param(model.index_of_node, model.time,initialize = 0, within = pe.Reals, mutable = True )
  for node in range (1,number_of_node+1):
    string_node = str(node).zfill(2) 
    #*Tạo số thứ tự có 2 chữ số (01, 02, ..., 36)
    
    filename = fr".\Node_PQ_data\MG{MG_index}\1Phase_24hLoad_MG{MG_index}{string_node}.xlsx"
    filepath = Path(filename)
    try:
     if filepath.is_file():
      df = pd.read_excel(filepath, engine='openpyxl')
      t_start = getattr(model, 't_start', 0)
      H = getattr(model, 'H', 24)
      for idx, t in df['hour'].items():
        actual_t = t - 1
        if t_start <= actual_t < t_start + H:
            model.p_node[node, actual_t - t_start] = float(df.at[idx, f' P1 (kW)'])*scale_power/S_base
            model.q_node[node, actual_t - t_start] = float(df.at[idx, f' Q1 (kvar)'])*scale_power/S_base
    except Exception as e:
     print(f"Lỗi khi đọc file: {e}")
  return
def create_price_from_grid(model,filename):
  model.Price_buy = pe.Param(model.time, initialize = 0, within =pe.Reals, mutable = True)
  model.Price_sell = pe.Param(model.time, initialize = 0, within =pe.Reals, mutable = True)  
  df = pd.read_excel(filename, engine = "openpyxl")
  t_start = getattr(model, 't_start', 0)
  H = getattr(model, 'H', 24)
  for index, value in df['Hour'].items():
    if pd.notna(value) and t_start <= value < t_start + H: 
      t_idx = value - t_start
      model.Price_buy[t_idx] = df.at[index, 'Price_Buy']
      model.Price_sell[t_idx] = df.at[index, 'Price_Sell']  
  return
def create_PV_para(model,filename):
  df =pd.read_excel(filename, engine= "openpyxl")
  model.temp_P_PV = pe.Param(model.node_has_PV, model.time,initialize = 0, within = pe.Reals, mutable =True)
  model.temp_Q_PV = pe.Param(model.node_has_PV, model.time,initialize = 0, within = pe.Reals, mutable =True)
  model.temp_S_PV = pe.Param(model.node_has_PV, model.time,initialize = 0, within = pe.Reals, mutable =True)
  t_start = getattr(model, 't_start', 0)
  H = getattr(model, 'H', 24)
  for index, value in df['Hour'].items():
    if pd.notna(value) and t_start <= value < t_start + H:
      temp_P_PV = df.at[index, 'P_PV']
      temp_Q_PV = df.at[index, 'Q_PV']
      temp_S_PV = sqrt(pow(temp_P_PV, 2) + pow(temp_Q_PV, 2))
      t_idx = value - t_start
      for node in model.node_has_PV:
        model.temp_P_PV[node, t_idx] = temp_P_PV*scale_power/S_base
        model.temp_Q_PV[node, t_idx] = temp_Q_PV*scale_power/S_base
        model.temp_S_PV[node, t_idx] = temp_S_PV*scale_power/S_base 
      # print(value,df.at[index, 'P_PV'],df.at[index, 'Q_PV'])
  model.P_PV_peak = pe.Param(model.node_has_PV, model.time, initialize=model.temp_P_PV)
  model.Q_PV = pe.Param(model.node_has_PV, model.time, initialize=model.temp_Q_PV)
  model.S_PV = pe.Param(model.node_has_PV, model.time, initialize=model.temp_S_PV)
  del model.temp_P_PV
  del model.temp_Q_PV
  del model.temp_S_PV
  return
def create_Wind_para(model,filename):
  df =pd.read_excel(filename, engine= "openpyxl")
  model.temp_P_Wind = pe.Param(model.node_has_Wind, model.time,initialize = 0, within = pe.Reals, mutable =True)
  model.temp_Q_Wind = pe.Param(model.node_has_Wind, model.time,initialize = 0, within = pe.Reals, mutable =True)
  model.temp_S_Wind = pe.Param(model.node_has_Wind, model.time,initialize = 0, within = pe.Reals, mutable =True)
  t_start = getattr(model, 't_start', 0)
  H = getattr(model, 'H', 24)
  for index, value in df['Hour'].items():
    if pd.notna(value) and t_start <= value < t_start + H:
      temp_P_Wind = df.at[index, 'P_Wind']
      temp_Q_Wind = df.at[index, 'Q_Wind']
      temp_S_Wind = sqrt(pow(temp_P_Wind, 2) + pow(temp_Q_Wind, 2))
      t_idx = value - t_start
      for node in model.node_has_Wind:
        model.temp_P_Wind[node, t_idx] = temp_P_Wind*scale_power/S_base
        model.temp_Q_Wind[node, t_idx] = temp_Q_Wind*scale_power/S_base
        model.temp_S_Wind[node, t_idx] = temp_S_Wind*scale_power/S_base
      # print(value,df.at[index, 'P_Wind'],df.at[index, 'Q_Wind'])
  model.P_Wind_peak = pe.Param(model.node_has_Wind, model.time, initialize=model.temp_P_Wind)
  model.Q_Wind = pe.Param(model.node_has_Wind, model.time, initialize=model.temp_Q_Wind)
  model.S_Wind = pe.Param(model.node_has_Wind, model.time, initialize=model.temp_S_Wind)
  del model.temp_P_Wind
  del model.temp_Q_Wind
  del model.temp_S_Wind
  return
def create_BESS_DG_para(model,filename):
  df = pd.read_excel(filename, engine= 'openpyxl')
  name = 'abc'
  para_dict = {}
  for index, name_para in df['Parameter'].items():
    value = df.at[index, 'Value' ]
    if pd.isna(value):
      name = str(df.at[index, 'Parameter'])
    if pd.notna(value):
      # print(name_para+'_'+ name)
      para_dict[name_para+'_'+ name] = float(value)
  # print(para_dict)
  for name, value in para_dict.items():
    setattr(model, name, pe.Param(initialize=value)) 
    # print(name, getattr(model, name).value)
  return
def create_location_source(model,list_PV, list_Wind, list_BESS, list_DG):
  # list_PV = {3,8,10,12,14,26,36}
  # list_Wind = {32}
  # list_BESS ={19}
  # list_DG
  type_of_source = {
    'PV',
    'Wind',
    'BESS', 
    'DG'
  }
  for name in type_of_source:
    source_list =locals() ["list_" + name]
    source_list = sorted(source_list)
    setattr(model, "node_has_"+ name, pe.Set(initialize=source_list, within=pe.NonNegativeIntegers))
  
  return
def Build_tree_form(number_of_nodee,model):#*ham de tao duong di cua luoi theo danh sach ke
  # Khai báo tập con child_of: mỗi node sẽ có một tập hợp các node con
  model.child_of = pe.Set(model.index_of_node, initialize=lambda m, node: [], ordered= True)
  model.child_of[0].add(0)
  model.child_of[0].add(1)
  for i in range(1,number_of_nodee+1):
    model.child_of[i].add(i)
  for parent, child in model.node_links:
      model.child_of[parent].add(child)
  return 
def create_trading_para(model,filename):
  df =pd.read_excel(filename, engine= "openpyxl")
  trading_list = []
  for index, value in df['MG_from'].items():
    trading_list.append(int(df.at[index, 'node_from']))
    
  trading_list = sorted(trading_list)
  setattr(model, "node_has_trade", pe.Set(initialize=trading_list, within=pe.NonNegativeIntegers))
  model.trading_cost = pe.Param(model.node_has_trade,model.time,within = pe.Reals, mutable = True)
  for node in model.node_has_trade:
    for time in model.time:
      model.trading_cost[node,time] = model.Price_sell[time] 
  model.P_form_another_MG = pe.Param(model.node_has_trade,model.time,initialize = 0, mutable =True)
  return
def create_phat_ADMM(model):
  model.phat_ADMM = pe.Param(model.node_has_trade,model.time, initialize = 0.05,mutable=True )
  return

#! KHAI BÁO BIẾN DÙNG TRONG MODEL
def Create_variable(model): 
  #! Các biến chung
  List_variable = {'P_line', 'Q_line', 'U_node', 'I_line'}
  for name in List_variable:
    setattr(model, name, pe.Var(model.index_of_node, model.time))

  #! RE
  List_PV_variable = {'P_PV'}
  for name in List_PV_variable:
    setattr(model, name, pe.Var(model.node_has_PV, model.time))
  List_Wind_variable = {'P_Wind'}
  for name in List_Wind_variable:
    setattr(model, name, pe.Var(model.node_has_Wind, model.time))
    
  #! Biến DG
  List_DG_variable_real = {'P_DG', 'On_cost_DG'}
  for name in List_DG_variable_real:
    setattr(model, name, pe.Var(model.node_has_DG, model.time, within=pe.NonNegativeReals))
  setattr(model, 'Q_DG', pe.Var(model.node_has_DG, model.time, within=pe.Reals))
  
  setattr(model, 'I_DG', pe.Var(model.node_has_DG, model.time, within=pe.Binary))

  List_variable_ontime = {'R_DG'}
  for name in List_variable_ontime:
    setattr(model, name, pe.Var(model.time))

  #! Biến BESS
  List_BESS_variable_real = {'PC_BESS', 'PD_BESS', 'SOC'}
  for name in List_BESS_variable_real:
    setattr(model, name, pe.Var(model.node_has_BESS, model.time))

  List_BESS_variable_binary = {'IC_BESS', 'ID_BESS'}
  for name in List_BESS_variable_binary:
    setattr(model, name, pe.Var(model.node_has_BESS, model.time, within=pe.Binary))

  #! Tham số ATC (Tie-line giữa DSO và MG)
  setattr(model, 'lamda_ATC', pe.Param(model.node_has_trade, model.time, mutable=True, initialize=0.0))
  setattr(model, 'rho_ATC', pe.Param(model.node_has_trade, model.time, mutable=True, initialize=0.0))
  setattr(model, 'P_target', pe.Param(model.node_has_trade, model.time, mutable=True, initialize=0.0))

  #! Biến Load Shedding
  setattr(model, 'P_shed', pe.Var(model.index_of_node, model.time, within=pe.NonNegativeReals))
  setattr(model, 'P_shed_normal', pe.Var(model.index_of_node, model.time, within=pe.NonNegativeReals))
  setattr(model, 'P_shed_critical', pe.Var(model.index_of_node, model.time, within=pe.NonNegativeReals))
  #! Biến Trade
  setattr(model, 'P_trade', pe.Var(model.node_has_trade, model.time))
  return 

#! OBJECTIVE FUNCTION 
def object_function(model, export_limits, stage='Normal'):
    lam = 1e8  # hệ số phạt
    fix_value = scale_power/S_base
    #* grid
    model.split_power = pe.ConstraintList()
    model.P_pos = pe.Var(model.time, within=pe.NonNegativeReals)
    model.P_neg = pe.Var(model.time, within=pe.NonPositiveReals)
    for t in model.time:
      model.split_power.add(model.P_line[0, t] == model.P_pos[t] + model.P_neg[t])
                
    model.selling_grid_limit = pe.ConstraintList()
    EXPORT_LIMIT = abs(export_limits)
    model.P_neg_paid = pe.Var(model.time, within=pe.NonPositiveReals) # Phần được trả tiền
    model.P_neg_free = pe.Var(model.time, within=pe.NonPositiveReals) # Phần bán free (vượt ngưỡng)  
    for t in model.time:
      # Logic 1: P_neg tổng phải bằng tổng 2 thành phần con
      model.selling_grid_limit.add(model.P_neg[t] == model.P_neg_paid[t] + model.P_neg_free[t])
      
      # Logic 2: Giới hạn phần được trả tiền (P_neg_paid >= -200)
      model.selling_grid_limit.add(model.P_neg_paid[t] >= -EXPORT_LIMIT)
    
    profit1_form_grid = sum(model.Price_buy[t]*model.P_pos[t]/fix_value + model.Price_sell[t]*model.P_neg[t]/fix_value
                 for t in model.time)
    
    #* DG
    profit2_form_DG = 0
    for time in model.time:
      for node in model.node_has_DG:
        profit2_form_DG += model.alpha_DG * model.I_DG[node,time] + model.beta_DG * model.P_DG[node,time]/fix_value+ model.theta_DG * model.P_DG[node,time]/fix_value* model.P_DG[node,time]/fix_value+ model.Cold_start_cost_DG*model.On_cost_DG[node,time]+ model.keta_DG * model.P_DG[node,time]/fix_value
        
    #* BESS   
    profit3_form_BESS= 0
    for time in model.time:
      for node in model.node_has_BESS:
        profit3_form_BESS += model.Theta_BESS *(model.Eff_BESS *model.PC_BESS[node,time]/fix_value+ model.PD_BESS[node,time]/fix_value/model.Eff_BESS) 
    
    #* Pelnaty for I_line
    penalty_current = 0
    for time in model.time:
      for node in model.index_of_node:
        if node > 0: 
          penalty_current += 0.1 * model.I_line[node, time]
          
    #* Penalty for ATC
    penalty_ATC = 0
    for time in model.time:
      for node in model.node_has_trade:
        p_tr_pu = model.P_trade[node, time]
        p_tg_pu = model.P_target[node, time]
        penalty_ATC += model.lamda_ATC[node, time] * (p_tr_pu / fix_value) + (model.rho_ATC[node, time] / 2) * (p_tr_pu - p_tg_pu)**2 / fix_value
          
    if stage == 'Normal':
        C_shed_normal = 250000
        C_shed_critical = 2500000
    elif stage == 'Emergency':
        # VOLL (Value of Lost Load) adjusted to avoid Economic Shedding trap
        # C_shed_normal > Diesel cost to force turning on DG instead of shedding
        C_shed_normal = 100000
        C_shed_critical = 5000000
        
    model.exp_without_trade = pe.Expression(expr=profit1_form_grid + profit2_form_DG + profit3_form_BESS + penalty_current)
    
    total_cost = profit1_form_grid + profit2_form_DG + profit3_form_BESS + penalty_current + penalty_ATC
    
    #* Tinh tong P_shed
    total_shed_normal = 0
    total_shed_critical = 0
    for time in model.time:
        for node in model.index_of_node:
            total_shed_normal += model.P_shed_normal[node, time]
            total_shed_critical += model.P_shed_critical[node, time]

    #* SOC penalty when using MPC
    SOC_DA_ref = getattr(model, 'SOC_DA_ref', None)
    H = getattr(model, 'H', 24)
    beta = 150000
    SOC_penalty = 0
    
    if SOC_DA_ref is not None:
        model.SOC_slack = pe.Var(model.node_has_BESS, within=pe.NonNegativeReals)
        model.SOC_slack_con = pe.ConstraintList()
        for node in model.node_has_BESS:
            model.SOC_slack_con.add(model.SOC_slack[node] >= SOC_DA_ref - model.SOC[node, H-1])
            SOC_penalty += beta * (model.SOC_slack[node] ** 2)

    model.OBJ = pe.Objective(expr= total_cost + C_shed_normal * total_shed_normal + C_shed_critical * total_shed_critical + SOC_penalty, sense=pe.minimize)
        
    return  

#! CONSTRAINTS 
#* System 
def hard_constraint_shed_critical(model):
  model.shed_critical_hard = pe.ConstraintList()
  for node in model.node_has_critload:
    for time in model.time:
      lhs = model.P_shed[node, time]
      rhs = 0
      model.shed_critical_hard.add(lhs == rhs)
  return

def limit_shed_to_load(model):
  model.limit_shed = pe.ConstraintList()
  for node in model.index_of_node:
    for time in model.time:
      # P_shed = P_shed_normal + P_shed_critical
      model.limit_shed.add(model.P_shed[node, time] == model.P_shed_normal[node, time] + model.P_shed_critical[node, time])
      # Limit P_shed_normal to 80% of node load
      model.limit_shed.add(model.P_shed_normal[node, time] <= 0.8 * model.p_node[node, time])
      # Limit P_shed_critical to 20% of node load
      model.limit_shed.add(model.P_shed_critical[node, time] <= 0.2 * model.p_node[node, time])
  return
def active_power_balance_constraints(model): #* 1a
  model.active_power_balance = pe.ConstraintList()

  for node in model.index_of_node: #*load
    # print(node,model.child_of[node].value)
    for time in model.time:
        P_lhs = model.P_line[node,time]- model.p_node[node,time] - model.R_line[node]*model.I_line[node, time] 
        sum_DG = 0
        sum_RG = 0
        sum_BESS = 0
        
        sum_Shed = model.P_shed[node,time]
        
        if node in model.node_has_DG:
          sum_DG += model.P_DG[node,time]
        if node in model.node_has_Wind:
          sum_RG += model.P_Wind[node,time]
        if node in model.node_has_PV:
          sum_RG += model.P_PV[node,time]
        if node in model.node_has_BESS:
          sum_BESS += model.PD_BESS[node,time] - model.PC_BESS[node,time]
        
        sum_Trade = 0
        if node in model.node_has_trade:
          sum_Trade += model.P_trade[node,time]
          
        rhs = 0
        for child_node in model.child_of[node]:
          if child_node != node:
            rhs += model.P_line[child_node, time]        
        model.active_power_balance.add( P_lhs + sum_DG + sum_RG + sum_BESS + sum_Shed + sum_Trade == rhs )
  return
def reactive_power_balance_constraints(model): #* 1b
  model.reactive_power_balance = pe.ConstraintList()
  
  for node in model.index_of_node:
    for time in model.time:
      q_shed_val = model.P_shed[node, time] * (model.q_node[node, time] / model.p_node[node, time]) if value(model.p_node[node, time]) > 0 else 0
      lhs = model.Q_line[node,time]- model.q_node[node,time] -model.X_line[node]* model.I_line[node, time] + q_shed_val
      sum_DG = 0
      if node in model.node_has_DG: sum_DG += model.Q_DG[node,time]
      sum_RG = 0
      sum_BESS = 0
      sum_Trade = 0
      rhs = 0
      for child_node in model.child_of[node]:
        if child_node != node:
          rhs += model.Q_line[child_node, time]
      model.reactive_power_balance.add( lhs +sum_DG+ sum_RG + sum_BESS == rhs )
  return
def limit_sell_to_grid(model,up,down):
  model.limit_buy_sell_form_grid = pe.ConstraintList()
  for time in model.time:
    P_lhs = model.P_line[0,time]
    # model.limit_buy_sell_form_grid.add( P_lhs<= up)
    model.limit_buy_sell_form_grid.add( P_lhs>= down)
  return
def voltage_balance_contraints(model):
  #* NOTE: các line ij ở đây đang được quy về lin trước của nó là P_line[i]
  model.voltage_balance = pe.ConstraintList()
  lhs = 0
  rhs = 0
  for time in model.time:
    for i in model.index_of_node:
      for j in model.child_of[i]:
        if i != j:
          lhs = model.U_node[i,time] - model.U_node[j,time]
          rhs = 2*(model.R_line[j]*model.P_line[j,time] + model.X_line[j]*model.Q_line[j,time]) 
          rhs -= (model.R_line[j]*model.R_line[j]+ model.X_line[j]*model.X_line[j])*model.I_line[j,time]
          model.voltage_balance.add(lhs == rhs)
  return
def slack_bus_constraints(model):
  model.slack_bus = pe.ConstraintList()
  for time in model.time:
    lhs = model.U_node[0,time]
    rhs = 1
    model.slack_bus.add(lhs == rhs)
  return
def voltage_limit_constraints(model, stage='Normal'):
  model.voltage_limit = pe.ConstraintList()
  for time in model.time:
    for node in model.index_of_node:
      lhs = model.U_node[node,time]
      rhs = 1.05*1.05
      model.voltage_limit.add(lhs <= rhs)
  for time in model.time:
    for node in model.index_of_node:
      lhs = model.U_node[node,time]
      if stage == 'Emergency':
          rhs = 0.90*0.90
      else:
          rhs = 0.95*0.95
      model.voltage_limit.add(lhs >= rhs)
  return
def socp_relaxation_constraints(model):
    model.socp_con = pe.ConstraintList()
    for time in model.time:
      for i in model.index_of_node:
        for j in model.child_of[i]:
          if i != j:
            lhs = (model.P_line[j, time]**2 + model.Q_line[j, time]**2)
            rhs = 1* model.I_line[j, time]
            model.socp_con.add(lhs <= rhs)
    return
def lb_limit_current(model,max_i):
  model.lb_limit_current = pe.ConstraintList()
  for time in model.time:
    for i in model.index_of_node:
          lhs = 0.001
          rhs = model.I_line[i, time]
          model.lb_limit_current.add(lhs <= rhs)
  for time in model.time:
    for i in model.index_of_node:
          lhs = model.I_line[i, time]
          rhs = max_i
          if i != 0 and i != 1:
            model.lb_limit_current.add(lhs <= rhs)
  return

#* RE
def active_power_PV_constraints(model): #*(4a)
  model.active_power_PV = pe.ConstraintList()
  for node in model.node_has_PV:
    for time in model.time:
      lhs = model.P_PV[node,time]
      rhs = model.P_PV_peak[node,time]
      model.active_power_PV.add(lhs <= rhs)
  for node in model.node_has_PV:
    for time in model.time:
      lhs = 0
      rhs = model.P_PV[node,time]
      model.active_power_PV.add(lhs <= rhs)
  return 
def active_power_Wind_constraints(model): #*(4b)
  model.active_power_Wind = pe.ConstraintList()
  for node in model.node_has_Wind:
    for time in model.time:
      lhs = model.P_Wind[node,time]
      rhs = model.P_Wind_peak[node,time]
      model.active_power_Wind.add(lhs <= rhs)
  for node in model.node_has_Wind:
    for time in model.time:
      lhs = 0
      rhs = model.P_Wind[node,time]
      model.active_power_Wind.add(lhs <= rhs)
  return

#* DG 
def active_power_DG_constraints(model): #*(4c)
  model.active_power_DG = pe.ConstraintList()
  for node in model.node_has_DG:
    for time in model.time:
      bound_down = model.P_Qmin_DG * model.I_DG[node,time]  #? cần hỏi
      valuee = model.P_DG[node,time]
      model.active_power_DG.add(bound_down <= valuee)
  for node in model.node_has_DG:
    for time in model.time:
      bound_up = model.P_Qmax_DG * model.I_DG[node,time]
      valuee = model.P_DG[node,time]
      model.active_power_DG.add(bound_up >= valuee)
  for node in model.node_has_DG:
    for time in model.time:
      S_max_val = pe.value(model.P_Qmax_DG) / 0.8
      lhs = model.P_DG[node, time]**2 + model.Q_DG[node, time]**2
      rhs = (S_max_val**2) * model.I_DG[node, time]
      model.active_power_DG.add(lhs <= rhs)
  return
def active_ramp_limit_constraints(model): #*(4ef)
  model.active_ramp_limit = pe.ConstraintList()
  for node in model.node_has_DG:
    for time in model.time:
      if (time +1) in model.time:
        lhs = model.P_DG[node,time+1] - model.P_DG[node,time]
        rhs = model.Ramp_rate_DG #? can hoi
        model.active_ramp_limit.add(lhs <= rhs)
  for node in model.node_has_DG:
    for time in model.time:
      if (time -1) in model.time:
        lhs = model.P_DG[node,time-1] - model.P_DG[node,time]
        rhs = model.Ramp_rate_DG
        model.active_ramp_limit.add(lhs <= rhs)
  return
def binary_DG_constraints(model):
  model.binary_DG = pe.ConstraintList()
  for node in model.node_has_DG:
    for time in model.time:
      lhs = 0
      rhs = model.I_DG[node,time]
      model.binary_DG.add(lhs<=rhs)
  for node in model.node_has_DG:
    for time in model.time:
      lhs = 1
      rhs = model.I_DG[node,time]
      model.binary_DG.add(lhs>=rhs)
  return
def Cost_on_constraints(model):
  model.Cost_on = pe.ConstraintList()
  for node in model.node_has_DG:
    for time in model.time:
      if time >0:
        lhs = model.On_cost_DG[node,time]
        rhs = model.I_DG[node,time]- model.I_DG[node,time-1]
        model.Cost_on.add(lhs>= rhs)
      elif time == 0:
        lhs = model.On_cost_DG[node,time]
        rhs = model.I_DG[node,0]
        model.Cost_on.add(lhs>= rhs)
  return
def up_time_constraint(model): #* (4m) 
  model.up_time = pe.ConstraintList()
  for node in model.node_has_DG:
    for time in model.time:
      lhs = 0 
      check = 0
      H = getattr(model, 'H', 24)
      for phi in range(int(time), int(time + model.Time_up_DG - 1)+1):
        if (int(time + model.Time_up_DG - 1)<= H - 1):
          lhs += model.I_DG[node,phi]
          check =1
      if (check ==1):
        rhs = model.Time_up_DG * model.On_cost_DG[node,time]
        model.up_time.add(lhs >= rhs)
  return
def down_time_constraint(model): #* (4n) 
  model.down_time = pe.ConstraintList()
  for node in model.node_has_DG:
    for time in model.time:
      lhs = 0 
      check = 0
      H = getattr(model, 'H', 24)
      for phi in range(int(time),int(time + model.Time_down_DG - 1) + 1):
        if (int(time + model.Time_down_DG - 1)<= H - 1):
          lhs += (1 - model.I_DG[node,phi])
          check = 1
      if (check == 1):
        if time > 0:
            rhs = model.Time_down_DG * (model.I_DG[node, time-1] - model.I_DG[node, time])
        else:
            rhs = 0
        model.down_time.add(lhs >= rhs)
  return
def Spinning_reserve_constraints(model): #*(4kl)
  model.Spinning_reserve = pe.ConstraintList()
  for time in model.time:
    lhs = model.R_DG[time]
    rhs = 0
    for node in model.node_has_DG:
      rhs += model.P_Qmax_DG * model.I_DG[node,time] - model.P_DG[node,time]
    model.Spinning_reserve.add(lhs == rhs)
  for time in model.time:
    lhs = model.R_DG[time]
    rhs = 0
    for node in model.node_has_DG:
      rhs += model.P_Qmax_DG
    model.Spinning_reserve.add(lhs >= 0.2 * rhs)
  return

#* BESS
def maximum_charging_power_constraints(model):#*(6ab)
  model.charging_power = pe.ConstraintList()
  for node in model.node_has_BESS:
    for time in model.time:
      lhs = model.PC_BESS[node,time]
      rhs = model.IC_BESS[node,time]* model.P_BESS_max_BESS 
      #? P_BESS_max PC_max
      model.charging_power.add(lhs <= rhs)
  for node in model.node_has_BESS:
    for time in model.time:
      lhs = 0
      rhs = model.PC_BESS[node,time]
      model.charging_power.add(lhs <= rhs) 
  return
def maximum_discharging_power_constraints(model):#*(6ac)
  model.discharging_power = pe.ConstraintList()
  for node in model.node_has_BESS:
    for time in model.time:
      lhs = model.PD_BESS[node,time]
      rhs = model.ID_BESS[node,time]* model.P_BESS_max_BESS 
      #? P_BESS_max PD_max
      model.discharging_power.add(lhs <= rhs)
  for node in model.node_has_BESS:
    for time in model.time:
      lhs = 0
      rhs = model.PD_BESS[node,time]
      model.discharging_power.add(lhs <= rhs) 
  return
def SOC_limit_constraints(model):#*(6d)
  model.SOC_limit = pe.ConstraintList()
  for node in model.node_has_BESS:
    for time in model.time:
      lhs = model.SOC_min_BESS
      rhs = model.SOC[node,time]
      model.SOC_limit.add(lhs<=rhs)
  for node in model.node_has_BESS:
    for time in model.time:
      rhs = model.SOC_max_BESS
      lhs = model.SOC[node,time]
      model.SOC_limit.add(lhs<=rhs)
  return
def SOC_temporal_variation_constraints(model): #*(6e)
  model.SOC_temporal_variation = pe.ConstraintList()
  SOC_init = getattr(model, 'SOC_init', None)
  for node in model.node_has_BESS:
    for time in model.time:
      if (time - 1)in model.time:
        lhs = model.SOC[node,time]
        rhs = model.SOC[node,time-1] + (model.PC_BESS[node,time]*model.Eff_BESS -model.PD_BESS[node,time]/model.Eff_BESS )/model.Capacity_BESS
        model.SOC_temporal_variation.add(lhs == rhs)
      else:
        lhs = model.SOC[node,time]
        rhs = SOC_init if SOC_init is not None else model.SOC_ini_BESS
        model.SOC_temporal_variation.add(lhs == rhs)
  for node in model.node_has_BESS:
    lhs = 0
    rhs = model.PC_BESS[node,0]
    model.SOC_temporal_variation.add(lhs == rhs)
    rhs = model.PD_BESS[node,0]
    model.SOC_temporal_variation.add(lhs == rhs)
  return
def SOC_equal_constraints(model): #*(6f)
  model.SOC_equal = pe.ConstraintList()
  SOC_init = getattr(model, 'SOC_init', None)
  H = getattr(model, 'H', 24)
  for node in model.node_has_BESS:
    day_begin = model.SOC[node,0]
    ini = SOC_init if SOC_init is not None else model.SOC_ini_BESS
    model.SOC_equal.add(day_begin == ini)
  if SOC_init is None or H == 24:
    for node in model.node_has_BESS:
      day_end = model.SOC[node, H-1]
      ini_end = SOC_init if SOC_init is not None else model.SOC_ini_BESS
      model.SOC_equal.add(day_end == ini_end)
  return
def BESS_I_limit_constraints(model): #*(6g)
  model.BESS_I_limit = pe.ConstraintList()
  for node in model.node_has_BESS:
    for time in model.time:
      lhs = model.IC_BESS[node,time] + model.ID_BESS[node,time]
      rhs = 1
      model.BESS_I_limit.add (lhs <= rhs)
  return

#* Trading
def Trading_limit_constraints(model,max_trading_limit,min_trading_limit):
  model.Trading_limit = pe.ConstraintList()
  for node in model.node_has_trade:
    for time in model.time:
      model.Trading_limit.add(model.P_trade[node, time] <= max_trading_limit)
      model.Trading_limit.add(model.P_trade[node, time] >= min_trading_limit)
  return
#! SOLVE
def solve(model):
  #!GLPK
  # path_glpk = "D:/Ungdung/Python/winglpk-4.65/glpk-4.65/w64/glpsol.exe"
  # solver = SolverFactory('glpk', executable= path_glpk)
  
  #!SCIP
  # path_scip = r"C:\Program Files\SCIPOptSuite 9.2.4\bin\scip.exe"
  # solver = SolverFactory('scip', executable=path_scip)
  
  #! IPOPT
  # path_ipopt = r"D:\Code\P2P_trading_withQV\my_env_P2PQV\Solver\ampl.mswin64\ipopt.exe"
  # solver = pe.SolverFactory('ipopt')

  #! GUROBI
  # solver = SolverFactory('gurobi', solver_io='python')
  #!MOSEK
  solver = SolverFactory('mosek')
  results = solver.solve(model, tee=False)
  # results = solver.solve(model, tee=True)
  
  return results

def print_active_shedding_nodes(model, mg_name):
    print(f"--- ACTIVE SHEDDING NODES FOR MG{mg_name} ---")
    shed_found = False
    for time in model.time:
        for node in model.index_of_node:
            shed_normal = pe.value(model.P_shed_normal[node, time])
            shed_critical = pe.value(model.P_shed_critical[node, time])
            
            if shed_normal > 1e-4 or shed_critical > 1e-4:
                print(f"[MG{mg_name}] Node: {node}, Time: {time} | Normal Shed: {shed_normal*1000:.2f} kW | Critical Shed: {shed_critical*1000:.2f} kW")
                shed_found = True
    if not shed_found:
        print(f"[MG{mg_name}] No load shedding active.")
    print("------------------------------------------")

def print_result(model,grid_index):
  return # Bỏ qua việc tạo file xlsx theo yêu cầu của user
  red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
  orange_fill = PatternFill(start_color="FABF8F", end_color="FABF8F", fill_type="solid")
  
  wb = Workbook()
  sheet = wb.active
  filename = f"result_of_MG{grid_index}"
  filepath = f".\\Result_data\\{filename}.xlsx"
  sheet.title = filename
  
  fix_value = scale_power/S_base
  #! IN RA P_line, Q_line
  sheet.cell(row = 1,column = 1).value= "Time(hour)"
  for time in model.time:
    sheet.cell(row = time+2 ,column = 1).value = time
  for node in model.index_of_node:
    sheet.cell(row = 1, column = node+2).value = f"S at node{node}"
    max_length = 0
    for time in model.time:
      P = round(model.P_line[node,time].value,4)
      Q = round(model.Q_line[node,time].value,4)
      sheet.cell(row = time+2, column = node+2).value = f"{P}+j{Q}"
      max_length = max(max_length,len(f"{P}+j{Q}"))
      # print(f"P_line[{node},{time}] = {model.P_line[node,time].value}")
    sheet.column_dimensions[get_column_letter(node+2)].width = len(f"{P}+j{Q}") + 2
 
  #! in ra cac var DG, P_DG,Q_DG,I_DG
  gian_dong =26
  gd = gian_dong
  tach_ngang = 0
  for time in model.time:
    sheet.cell(row = time+2+gd ,column = 1).value = time
  tach_ngang+= 2
  for node in model.node_has_DG:
    sheet.cell(row = 1+gd, column = tach_ngang).value = f"P_DG at node:{node}"
    sheet.cell(row = 1+gd, column = tach_ngang+1).value = f"Q_DG at node:{node}"
    sheet.cell(row = 1+gd, column = tach_ngang+2).value = f"I_DG at node:{node}"
    sheet.cell(row = 1+gd, column = tach_ngang+3).value = f"ON_DG at node:{node}"
    max_length = 0
    for time in model.time:
      P = model.P_DG[node,time].value
      P = P/ fix_value
      # Q = model.Q_DG[node,time].value
      Q = 0 
      I = model.I_DG[node,time].value
      ON = model.On_cost_DG[node,time].value
      if model.P_DG[node,time].value is not None:
        P= round (P,4)
        Q= round (Q,4)
        I= round (I,4)
      else:
        P = -36
        Q = -36
        I =-36
        ON = -36
      sheet.cell(row = time+2+gd, column = tach_ngang).value = P
      sheet.cell(row = time+2+gd, column = tach_ngang+1).value = Q
      sheet.cell(row = time+2+gd, column = tach_ngang+2).value = I
      sheet.cell(row = time+2+gd, column = tach_ngang+3).value = ON
    tach_ngang+=4

  #! in ra cac var BESS: PC_BESS,PD_BESS,IC_BESS,ID_BESS,SOC
  for node in model.node_has_BESS:
    sheet.cell(row = 1+gd, column = tach_ngang).value = f"PC_BESS at node:{node}"
    sheet.cell(row = 1+gd, column = tach_ngang+1).value = f"PD_BESS at node:{node}"
    sheet.cell(row = 1+gd, column = tach_ngang+2).value = f"IC_BESS at node:{node}"
    sheet.cell(row = 1+gd, column = tach_ngang+3).value = f"ID_BESS at node:{node}"
    sheet.cell(row = 1+gd, column = tach_ngang+4).value = f"SOC at node:{node}"
    max_length = 0
    for time in model.time:
      PC_BESS = model.PC_BESS[node, time].value
      PD_BESS = model.PD_BESS[node, time].value
      IC_BESS = model.IC_BESS[node, time].value
      ID_BESS = model.ID_BESS[node, time].value
      SOC     = model.SOC[node, time].value
      if model.PC_BESS[node, time].value is not None:
        PC_BESS = round(model.PC_BESS[node, time].value, 4)

      if model.PD_BESS[node, time].value is not None:
        PD_BESS = round(model.PD_BESS[node, time].value, 4)

      if model.IC_BESS[node, time].value is not None:
        IC_BESS = round(model.IC_BESS[node, time].value, 4)

      if model.ID_BESS[node, time].value is not None:
        ID_BESS = round(model.ID_BESS[node, time].value, 4)

      if model.SOC[node, time].value is not None:
        SOC = round(model.SOC[node, time].value, 4)
      PC_BESS = PC_BESS/ fix_value
      PD_BESS = PD_BESS/ fix_value
      sheet.cell(row=time+2+gd, column=tach_ngang).value = PC_BESS
      sheet.cell(row=time+2+gd, column=tach_ngang+1).value = PD_BESS
      sheet.cell(row=time+2+gd, column=tach_ngang+2).value = IC_BESS
      sheet.cell(row=time+2+gd, column=tach_ngang+3).value = ID_BESS
      sheet.cell(row=time+2+gd, column=tach_ngang+4).value = SOC
    tach_ngang+= 5
  #! in ra U_node
  gian_dong +=26
  gd = gian_dong
  for time in model.time:
    sheet.cell(row = time+2+gd ,column = 1).value = time
  for node in model.index_of_node:
    sheet.cell(row = 1+gd, column = node+2).value = f"U at node{node}"
    max_length = 0
    for time in model.time:
      U = model.U_node[node,time].value
      if model.U_node[node, time].value is not None:
        if U>0:
          U = sqrt(U)
        U= round (U,4)
      current_cell = sheet.cell(row=time + 2 + gd, column=node + 2)
      current_cell.value = U
      if U > 1.05:
        current_cell.fill = red_fill      # Tô nền màu đỏ nhạt
      if U < 0.95:
        current_cell.fill = orange_fill  # Tô nền màu cam
      if U < 0.9:
        current_cell.fill = red_fill  
  #! in ra I_line
  gian_dong +=26
  gd = gian_dong
  for time in model.time:
    sheet.cell(row = time+2+gd ,column = 1).value = time
  for node in model.index_of_node:
    sheet.cell(row = 1+gd, column = node+2).value = f"I at line{node}"
    max_length = 0
    for time in model.time:
      if node == 0:
        continue
      I = model.I_line[node,time].value
      if model.I_line[node, time].value is not None:
        if I > 0:
          I = sqrt(I)
        else:
          I = -sqrt(-I)
        I= round (I,4)
      current_cell = sheet.cell(row = time+2+gd, column = node+2)
      sheet.cell(row = time+2+gd, column = node+2).value = I
      if I < 0:
        current_cell.fill = red_fill      # Tô nền màu đỏ nhạt
      
  sheet.insert_rows(1)
  obj_expr = model.OBJ.expr
  
  sheet.cell(row =1, column= 1, value= "Ω=")
  sheet.cell(row =1, column= 2).value = obj_expr() 
  #! chot file
  wb.save(filepath)

      
  return
def create_graph_separate_MG_trading(model, exportlimit, trading_line, MG_index, model_name="default", model_da=None):
    t_start = getattr(model, 't_start', 0)
    H = getattr(model, 'H', len(model.time))
    total_len = t_start + H if model_da is not None else len(model.time)
    
    sum_P_node = [0] * total_len
    sum_P_shed = [0] * total_len
    sum_Wind = [0] * total_len
    sum_PV = [0] * total_len
    sum_BESS = [0] * total_len
    sum_P_line_0 = [0] * total_len
    sum_DG = [0] * total_len
    sum_Trade = np.zeros((5, total_len))
    
    #! --- 1. THU THẬP DỮ LIỆU ---
    if model_da is not None:
        for t in range(t_start):
            for node in model_da.index_of_node:
                if node in model_da.node_has_DG:
                    sum_DG[t] += model_da.P_DG[node, t].value
                if node in model_da.node_has_BESS:
                    sum_BESS[t] += -model_da.PC_BESS[node, t].value + model_da.PD_BESS[node, t].value
                if node in model_da.node_has_Wind:
                    sum_Wind[t] += model_da.P_Wind[node, t].value
                if node in model_da.node_has_PV:
                    sum_PV[t] += model_da.P_PV[node, t].value
                if node in model_da.node_has_trade:
                    MG_to = trading_line[MG_index, node]
                    sum_Trade[MG_to, t] += model_da.P_trade[node, t].value
                sum_P_node[t] += model_da.p_node[node, t].value
                sum_P_shed[t] += model_da.P_shed[node, t].value
            sum_P_line_0[t] += model_da.P_line[0, t].value

    for t_idx in model.time:
        t_real = t_start + t_idx if model_da is not None else t_idx
        for node in model.index_of_node:
            if node in model.node_has_DG:
                sum_DG[t_real] += model.P_DG[node, t_idx].value
            if node in model.node_has_BESS:
                sum_BESS[t_real] += -model.PC_BESS[node, t_idx].value + model.PD_BESS[node, t_idx].value
            if node in model.node_has_Wind:
                sum_Wind[t_real] += model.P_Wind[node, t_idx].value
            if node in model.node_has_PV:
                sum_PV[t_real] += model.P_PV[node, t_idx].value
            if node in model.node_has_trade:
                MG_to = trading_line[MG_index, node]
                sum_Trade[MG_to, t_real] += model.P_trade[node, t_idx].value
            sum_P_node[t_real] += model.p_node[node, t_idx].value
            sum_P_shed[t_real] += model.P_shed[node, t_idx].value
        sum_P_line_0[t_real] += model.P_line[0, t_idx].value
            
    #! --- 2. VẼ BIỂU ĐỒ ---
    fig, ax = plt.subplots(figsize=(10, 4))
    width = 0.8
    x = np.arange(total_len)
    
    # Vẽ đường Tải (Load) riêng biệt để so sánh
    # plt.bar(x, sum_P_node, width, color='skyblue', label='P_node (Load)', alpha=0.3) 
    plt.plot(x, sum_P_node, color='skyblue', marker='*', linestyle='-', linewidth=2, label='P_node (Load)', alpha=0.8)
    
    load_after_shed = [sum_P_node[i] - sum_P_shed[i] for i in range(total_len)]
    plt.plot(x, load_after_shed, color='red', linestyle='--', linewidth=2, label='Load after shed', alpha=0.8)
    
    # Chuẩn bị dữ liệu
    data = {
        'sum_DG': sum_DG,
        'sum_BESS': sum_BESS,
        'sum_P_line_0': sum_P_line_0,
        'sum_Wind': sum_Wind,
        'sum_PV': sum_PV,
        'Trade_MG0': sum_Trade[0], 
        'Trade_MG1': sum_Trade[1], 
        'Trade_MG2': sum_Trade[2],
        'Trade_MG3': sum_Trade[3],
        'Trade_MG4': sum_Trade[4]
    }

    classified_data = {}
    for name, arr in data.items():
        classified_data[name + '_pos'] = np.clip(arr, 0, None)
        classified_data[name + '_neg'] = np.clip(arr, None, 0)

    bottom_pos = np.zeros_like(list(data.values())[0],dtype = float)
    bottom_neg = np.zeros_like(list(data.values())[0],dtype = float)

    # --- THAY ĐỔI QUAN TRỌNG Ở ĐÂY ---
    # Đưa các thành phần Trade lên đầu danh sách để vẽ trước -> Nằm sát trục 0
    sources = [
         'sum_P_line_0','Trade_MG0', 'Trade_MG1', 'Trade_MG2', 'Trade_MG3', 'Trade_MG4',  # Vẽ trước (sát 0)
        'sum_DG', 'sum_BESS', 'sum_Wind', 'sum_PV' # Vẽ sau (chồng lên trên)
    ]
    colors = [
        'purple','yellow', 'brown', 'pink', 'blue', 'gray',   # Màu cho Trade
        'red', 'cyan', 'green', 'orange' # Màu cho các nguồn khác
    ]   
    labels = [
        'P_Grid','Trade w/ MG0', 'Trade w/ MG1', 'Trade w/ MG2', 'Trade w/ MG3', 'Trade w/ MG4',
        'P_DG','P_BESS', 'P_Wind', 'P_PV'
    ]
    if total_len > 16:
        sum_DG[16] = 0
    # Vòng lặp vẽ
    for i, source in enumerate(sources):
        pos_data = classified_data[source + '_pos']
        neg_data = classified_data[source + '_neg']
        # if(source == 'Trade_MG1' or source == 'Trade_MG2' or source == 'Trade_MG3' or source == 'Trade_MG4'):
        #   print(source,f"sum_Trade pos:{pos_data}\n",f"sum_Trade neg:{neg_data}")
          # Chỉ vẽ nếu có dữ liệu khác 0
        if np.any(pos_data) or np.any(neg_data): 
            # Vẽ phần dương
            ax.bar(x, pos_data, width, bottom=bottom_pos, color=colors[i], label=labels[i])
            bottom_pos += pos_data

            # Vẽ phần âm
            ax.bar(x, neg_data, width, bottom=bottom_neg, color=colors[i])
            bottom_neg += neg_data
    # plt.axhline(-exportlimit*scale_power/S_base, color='red', linestyle='--', linewidth=1.5, label='Export Limit')
    # Cài đặt hiển thị
    # ax.set_title(f'Biểu đồ cân bằng năng lượng của MG {MG_index}')
    ax.set_xlabel('Time (h)')
    ax.set_ylabel('Energy (MW)')
    
    # Legend
    ax.legend(loc='upper left', ncol=2, fancybox=True, shadow=False)
    
    plt.axhline(0, color='black', linewidth=0.8)
    plt.tight_layout()
    
    import os
    save_dir = f'.\\Result_data\\report_result\\stage_2'
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    filename = f"{model_name}_Active_Power_MG{MG_index}"
    filepath = f"{save_dir}\\{filename}.png"
    plt.savefig(filepath, dpi=300, bbox_inches='tight')
    # plt.show()
    plt.close(fig)

def create_graph_separate_MG_Q(model, exportlimit, trading_line, MG_index, model_name="default", model_da=None):
    import pyomo.environ as pe
    import numpy as np
    import matplotlib.pyplot as plt
    import math

    t_start = getattr(model, 't_start', 0)
    H = getattr(model, 'H', len(model.time))
    total_len = t_start + H if model_da is not None else len(model.time)
    
    sum_q_node = [0] * total_len
    sum_q_shed = [0] * total_len
    sum_p_node = [0] * total_len
    sum_p_shed = [0] * total_len
    sum_Q_line_0 = [0] * total_len
    sum_Q_DG = [0] * total_len
    
    #! --- 1. THU THẬP DỮ LIỆU ---
    if model_da is not None:
        for t in range(t_start):
            for node in model_da.index_of_node:
                p_n = pe.value(model_da.p_node[node, t])
                q_n = pe.value(model_da.q_node[node, t])
                p_sh = pe.value(model_da.P_shed[node, t])
                q_sh = p_sh * (q_n / p_n) if p_n > 0 else 0
                
                sum_p_node[t] += p_n
                sum_p_shed[t] += p_sh
                sum_q_node[t] += q_n
                sum_q_shed[t] += q_sh
                
                if node in model_da.node_has_DG:
                    sum_Q_DG[t] += pe.value(model_da.Q_DG[node, t])
            sum_Q_line_0[t] += pe.value(model_da.Q_line[0, t])

    for t_idx in model.time:
        t_real = t_start + t_idx if model_da is not None else t_idx
        for node in model.index_of_node:
            p_n = pe.value(model.p_node[node, t_idx])
            q_n = pe.value(model.q_node[node, t_idx])
            p_sh = pe.value(model.P_shed[node, t_idx])
            q_sh = p_sh * (q_n / p_n) if p_n > 0 else 0
            
            sum_p_node[t_real] += p_n
            sum_p_shed[t_real] += p_sh
            sum_q_node[t_real] += q_n
            sum_q_shed[t_real] += q_sh
            
            if node in model.node_has_DG:
                sum_Q_DG[t_real] += pe.value(model.Q_DG[node, t_idx])
        sum_Q_line_0[t_real] += pe.value(model.Q_line[0, t_idx])
            
    # Tính toán Hourly Power Factor
    hourly_pf = []
    for t in range(total_len):
        p_t = sum_p_node[t] - sum_p_shed[t]
        q_t = sum_q_node[t] - sum_q_shed[t]
        if p_t > 0:
            pf_t = p_t / math.sqrt(p_t**2 + q_t**2)
            hourly_pf.append(pf_t)

    if hourly_pf:
        min_pf = min(hourly_pf)
        max_pf = max(hourly_pf)
        avg_pf = sum(hourly_pf) / len(hourly_pf)
    else:
        min_pf = max_pf = avg_pf = 0
    
    print(f"[MG{MG_index}] Hourly PF -> Min: {min_pf:.4f}, Max: {max_pf:.4f}, Avg: {avg_pf:.4f}")
    with open(f".\\Result_data\\PF_MG{MG_index}.txt", "w") as f:
        f.write(f"MG{MG_index} Power Factor Statistics:\n")
        f.write(f"Min PF: {min_pf:.4f}\n")
        f.write(f"Max PF: {max_pf:.4f}\n")
        f.write(f"Avg PF: {avg_pf:.4f}\n")
        f.write(f"Total P Load (after shed): {sum(sum_p_node) - sum(sum_p_shed):.4f} MW\n")
        f.write(f"Total Q Load (after shed): {sum(sum_q_node) - sum(sum_q_shed):.4f} MVAR\n\n")
        f.write("Hourly PF:\n")
        for t in range(total_len):
            p_t = sum_p_node[t] - sum_p_shed[t]
            q_t = sum_q_node[t] - sum_q_shed[t]
            pf_t = p_t / math.sqrt(p_t**2 + q_t**2) if p_t > 0 else 0
            f.write(f"Hour {t}: {pf_t:.4f}\n")

    #! --- 2. VẼ BIỂU ĐỒ ---
    fig, ax = plt.subplots(figsize=(10, 4))
    width = 0.8
    x = np.arange(total_len)
    
    plt.plot(x, sum_q_node, color='skyblue', marker='*', linestyle='-', linewidth=2, label='Q_node (Load)', alpha=0.8)
    
    q_load_after_shed = [sum_q_node[i] - sum_q_shed[i] for i in range(total_len)]
    plt.plot(x, q_load_after_shed, color='red', linestyle='--', linewidth=2, label='Q Load after shed', alpha=0.8)

    data = {
        'sum_Q_DG': sum_Q_DG,
        'sum_Q_line_0': sum_Q_line_0
    }

    classified_data = {}
    for name, arr in data.items():
        classified_data[name + '_pos'] = np.clip(arr, 0, None)
        classified_data[name + '_neg'] = np.clip(arr, None, 0)

    bottom_pos = np.zeros(total_len)
    bottom_neg = np.zeros(total_len)

    sources = ['sum_Q_line_0', 'sum_Q_DG']
    colors = ['purple', 'red']   
    labels = ['Q_Grid', 'Q_DG']

    for i, source in enumerate(sources):
        pos_data = classified_data[source + '_pos']
        neg_data = classified_data[source + '_neg']
        if np.any(pos_data) or np.any(neg_data): 
            ax.bar(x, pos_data, width, bottom=bottom_pos, color=colors[i], label=labels[i])
            bottom_pos += pos_data
            ax.bar(x, neg_data, width, bottom=bottom_neg, color=colors[i])
            bottom_neg += neg_data

    ax.set_title(f'MG{MG_index} Q Balance (PF: Min={min_pf:.4f}, Max={max_pf:.4f}, Avg={avg_pf:.4f})')
    ax.set_xlabel('Time (h)')
    ax.set_ylabel('Reactive Power (MVAR)')
    ax.legend(loc='upper left', ncol=2, fancybox=True, shadow=False)
    plt.axhline(0, color='black', linewidth=0.8)
    plt.tight_layout()
    
    import os
    save_dir = f'.\\Result_data\\report_result\\stage_3'
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    filename = f"{model_name}_Reactive_Power_MG{MG_index}"
    filepath = f"{save_dir}\\{filename}.png"
    plt.savefig(filepath, dpi=300, bbox_inches='tight')
    plt.close(fig)
    # Add call to generate IEEE standard graphs
    plot_ieee_graphs(model, exportlimit, trading_line, MG_index, model_name=model_name, model_da=model_da)

    return

def plot_ieee_graphs(model, exportlimit, trading_line, MG_index, model_name="default", model_da=None):
    import os
    if not os.path.exists(f'.\\Result_data\\{model_name}'):
        os.makedirs(f'.\\Result_data\\{model_name}')
    
    t_start = getattr(model, 't_start', 0)
    H = getattr(model, 'H', len(model.time))
    total_len = t_start + H if model_da is not None else len(model.time)
    
    trade_data = {}
    soc_data = {}
    shed_normal_nodes = {}
    shed_critical_nodes = {}
    
    # Extract data
    if model_da is not None:
        for t in range(t_start):
            for node in model_da.index_of_node:
                if node in model_da.node_has_trade:
                    MG_to_raw = trading_line[MG_index, node]
                    MG_to = MG_to_raw[0] if isinstance(MG_to_raw, list) else MG_to_raw
                    if MG_to not in trade_data: trade_data[MG_to] = [0] * total_len
                    trade_data[MG_to][t] += value(model_da.P_trade[node, t])
                if node in model_da.node_has_BESS:
                    if node not in soc_data: soc_data[node] = [0] * total_len
                    soc_data[node][t] = value(model_da.SOC[node, t])
                if hasattr(model_da, 'P_shed_normal'):
                    val_sn = value(model_da.P_shed_normal[node, t])
                    if val_sn > 1e-4:
                        if node not in shed_normal_nodes: shed_normal_nodes[node] = [0]*total_len
                        shed_normal_nodes[node][t] += val_sn
                if hasattr(model_da, 'P_shed_critical'):
                    val_sc = value(model_da.P_shed_critical[node, t])
                    if val_sc > 1e-4:
                        if node not in shed_critical_nodes: shed_critical_nodes[node] = [0]*total_len
                        shed_critical_nodes[node][t] += val_sc
                elif hasattr(model_da, 'P_shed'):
                    val_s = value(model_da.P_shed[node, t])
                    if val_s > 1e-4:
                        if node not in shed_normal_nodes: shed_normal_nodes[node] = [0]*total_len
                        shed_normal_nodes[node][t] += val_s

    for t_idx in model.time:
        t_real = t_start + t_idx if model_da is not None else t_idx
        for node in model.index_of_node:
            if node in model.node_has_trade:
                MG_to_raw = trading_line[MG_index, node]
                MG_to = MG_to_raw[0] if isinstance(MG_to_raw, list) else MG_to_raw
                if MG_to not in trade_data: trade_data[MG_to] = [0] * total_len
                trade_data[MG_to][t_real] += value(model.P_trade[node, t_idx])
            if node in model.node_has_BESS:
                if node not in soc_data: soc_data[node] = [0] * total_len
                soc_data[node][t_real] = value(model.SOC[node, t_idx])
            if hasattr(model, 'P_shed_normal'):
                val_sn = value(model.P_shed_normal[node, t_idx])
                if val_sn > 1e-4:
                    if node not in shed_normal_nodes: shed_normal_nodes[node] = [0]*total_len
                    shed_normal_nodes[node][t_real] += val_sn
            if hasattr(model, 'P_shed_critical'):
                val_sc = value(model.P_shed_critical[node, t_idx])
                if val_sc > 1e-4:
                    if node not in shed_critical_nodes: shed_critical_nodes[node] = [0]*total_len
                    shed_critical_nodes[node][t_real] += val_sc
            elif hasattr(model, 'P_shed'):
                val_s = value(model.P_shed[node, t_idx])
                if val_s > 1e-4:
                    if node not in shed_normal_nodes: shed_normal_nodes[node] = [0]*total_len
                    shed_normal_nodes[node][t_real] += val_s

    # 1. Plot P2P Power Flow
    fig1, ax1 = plt.subplots(figsize=(8, 4))
    x = np.arange(total_len)
    markers = ['o', 's', '^', 'D', 'v', '<', '>']
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2']
    plot_trade = False
    for i, (mg_to, data) in enumerate(trade_data.items()):
        if any(abs(v) > 1e-4 for v in data):
            ax1.plot(x, data, marker=markers[i%len(markers)], color=colors[i%len(colors)], 
                    linestyle='-', linewidth=1.5, label=f'To MG {mg_to}')
            plot_trade = True
    if plot_trade:
        ax1.set_title(f'P2P Power Flow from MG {MG_index}', fontsize=12, fontweight='bold')
        ax1.set_xlabel('Time (h)', fontsize=10)
        ax1.set_ylabel('Power Trade (MW)', fontsize=10)
        ax1.grid(True, linestyle='--', alpha=0.7)
        ax1.legend(loc='best', fontsize=9)
        fig1.tight_layout()
        fig1.savefig(f".\\Result_data\\{model_name}\\P2P_PowerFlow_MG{MG_index}.png", dpi=300, bbox_inches='tight')
    plt.close(fig1)

    # 2. Plot BESS SOC
    fig2, ax2 = plt.subplots(figsize=(8, 4))
    plot_soc = False
    for i, (node, data) in enumerate(soc_data.items()):
        if any(abs(v) > 1e-4 for v in data):
            ax2.plot(x, data, marker=markers[i%len(markers)], color=colors[i%len(colors)], 
                    linestyle='-', linewidth=1.5, label=f'Node {node}')
            plot_soc = True
    if plot_soc:
        ax2.set_title(f'BESS State of Charge (SOC) - MG {MG_index}', fontsize=12, fontweight='bold')
        ax2.set_xlabel('Time (h)', fontsize=10)
        ax2.set_ylabel('SOC', fontsize=10)
        ax2.grid(True, linestyle='--', alpha=0.7)
        if hasattr(model, 'SOC_min_BESS'):
            ax2.axhline(value(model.SOC_min_BESS), color='red', linestyle='--', alpha=0.6, label='Min SOC')
        if hasattr(model, 'SOC_max_BESS'):
            ax2.axhline(value(model.SOC_max_BESS), color='green', linestyle='--', alpha=0.6, label='Max SOC')
        ax2.legend(loc='best', fontsize=9)
        fig2.tight_layout()
        fig2.savefig(f".\\Result_data\\{model_name}\\BESS_SOC_MG{MG_index}.png", dpi=300, bbox_inches='tight')
    plt.close(fig2)

    # 3. Plot Load Shedding
    fig3, ax3 = plt.subplots(figsize=(8, 4))
    has_shedding = False
    bottom_normal = np.zeros(total_len)
    bottom_critical = np.zeros(total_len)
    
    colors = plt.rcParams['axes.prop_cycle'].by_key()['color']
    
    for idx, node in enumerate(shed_normal_nodes.keys()):
        data = np.array(shed_normal_nodes[node])
        ax3.bar(x, data, width=0.6, bottom=bottom_normal, label=f'Node {node} (Normal)', color=colors[idx % len(colors)])
        bottom_normal += data
        has_shedding = True
        
    for idx, node in enumerate(shed_critical_nodes.keys()):
        data = np.array(shed_critical_nodes[node])
        bottom_critical_plot = bottom_normal + bottom_critical
        ax3.bar(x, data, width=0.6, bottom=bottom_critical_plot, label=f'Node {node} (Critical)', hatch='//', color=colors[(idx + 4) % len(colors)])
        bottom_critical += data
        has_shedding = True

    if has_shedding:
        ax3.set_title(f'Load Shedding by Node - MG {MG_index}', fontsize=12, fontweight='bold')
        ax3.set_xlabel('Time (h)', fontsize=10)
        ax3.set_ylabel('Power Shed (MW)', fontsize=10)
        ax3.grid(True, axis='y', linestyle='--', alpha=0.7)
        ax3.legend(loc='center left', bbox_to_anchor=(1, 0.5), fontsize=9)
        fig3.tight_layout()
        fig3.savefig(f".\\Result_data\\{model_name}\\Load_Shedding_MG{MG_index}.png", dpi=300, bbox_inches='tight')
    plt.close(fig3)
def print_line_data(model, trading_line, MG_index):
  for time in model.time:
      print(f"Time: {time}")
      for node in model.index_of_node:       
          print (f"node {node}: P_node: {model.p_node[node, time].value}, P_line: {model.P_line[node, time].value}")
          if node > 0:
              print(f"Loss: {value(model.R_line[node]*model.I_line[node, time])}")
          if node in model.node_has_DG:
              print(f"P_DG: {model.P_DG[node, time].value}")
          if node in model.node_has_BESS:
              print(f"P_BESS: {-model.PC_BESS[node, time].value + model.PD_BESS[node, time].value}")
          if node in model.node_has_Wind:
              print(f"P_Wind: {model.P_Wind[node, time].value}")
          if node in model.node_has_PV:
              print(f"P_PV: {model.P_PV[node, time].value}")
          if node in model.node_has_trade:
              MG_to = trading_line[MG_index, node]
              print(f"P_trade to MG{MG_to}: {model.P_trade[node, time].value}")           
  return
def print_total_loss_data(model, trading_line, MG_index):
  source = [0]*24
  loss = [0]*24
  load = [0]*24
  for time in model.time:
      # print(f"Time: {time}")
      for node in model.index_of_node:
          # print (f"node {node}: P_node: {model.p_node[node, time].value}, voltage: {model.U_node[node, time].value}, P_line: {model.P_line[node, time].value}")
          load[time] += model.p_node[node, time].value   
          if node > 0:
              loss[time] += value(model.R_line[node]*model.I_line[node, time])
              # print(f"Loss: {value(model.R_line[node]*model.I_line[node, time])}")
              # print(f"node:{node}, voltage: {model.U_node[node, time].value},I_line: {sqrt(model.I_line[node, time].value)}, I_over:{value(model.I_line[node, time].value- model.P_line[node, time]**2 - model.Q_line[node, time]**2)}")
          if node in model.node_has_DG:
              source[time] += model.P_DG[node, time].value
              # print(f"P_DG: {model.P_DG[node, time].value}")
          if node in model.node_has_BESS:
              source[time] += -model.PC_BESS[node, time].value + model.PD_BESS[node, time].value
              # print(f"P_BESS: {-model.PC_BESS[node, time].value + model.PD_BESS[node, time].value}")
          if node in model.node_has_Wind:
              source[time] += model.P_Wind[node, time].value 
              # print(f"P_Wind: {model.P_Wind[node, time].value}")           
          if node in model.node_has_PV:
              source[time] += model.P_PV[node, time].value   
              # print(f"P_PV: {model.P_PV[node, time].value}")           
          if node in model.node_has_trade:
              MG_to = trading_line[MG_index, node]
              source[time] += model.P_trade[node, time].value
              # print(f"P_trade to MG{MG_to}: {model.P_trade[node, time].value}")
          if node == 0:
            source[time] += model.P_line[0, time].value
    
  # for t in range(24):
  #     print(f"{t}\t{load[t]}\t{source[t]}\t{loss[t]}")
  return load, loss 
def compare_re_curtailment(model_MG, model_MG_base):
    """
    Analyzes and plots the Renewable Energy (RE) curtailment comparison.
    - Base Case (Independent): Stacked Area Chart
    - Trading Case (P2P): Overlay Line Chart
    Prints a direct comparison of utilized RE.
    """
    hours = np.arange(24)

    # ==========================================
    # PART 1: CALCULATE FOR BASE CASE (model_MG_base)
    # ==========================================
    data_to_plot_base = {mg: np.zeros(24) for mg in model_MG_base}

    for mg in model_MG_base:
        for node in model_MG_base[mg].node_has_PV:
            for t in model_MG_base[mg].time:
                peak = model_MG_base[mg].P_PV_peak[node, t]
                actual = model_MG_base[mg].P_PV[node, t].value
                diff = max(0, peak - actual)
                if 0 <= t < 24: data_to_plot_base[mg][t] += diff
                    
        for node in model_MG_base[mg].node_has_Wind:
            for t in model_MG_base[mg].time:
                peak = model_MG_base[mg].P_Wind_peak[node, t]
                actual = model_MG_base[mg].P_Wind[node, t].value
                diff = max(0, peak - actual)
                if 0 <= t < 24: data_to_plot_base[mg][t] += diff

    # ==========================================
    # PART 2: CALCULATE FOR TRADING CASE (model_MG)
    # ==========================================
    data_to_plot_trading = {mg: np.zeros(24) for mg in model_MG}

    for mg in model_MG:
        for node in model_MG[mg].node_has_PV:
            for t in model_MG[mg].time:
                peak = model_MG[mg].P_PV_peak[node, t]
                actual = model_MG[mg].P_PV[node, t].value
                diff = max(0, peak - actual)
                if 0 <= t < 24: data_to_plot_trading[mg][t] += diff
                    
        for node in model_MG[mg].node_has_Wind:
            for t in model_MG[mg].time:
                peak = model_MG[mg].P_Wind_peak[node, t]
                actual = model_MG[mg].P_Wind[node, t].value
                diff = max(0, peak - actual)
                if 0 <= t < 24: data_to_plot_trading[mg][t] += diff

    # ==========================================
    # PART 3: DATA PROCESSING FOR PLOTTING
    # ==========================================
    y_values_base = [data_to_plot_base[mg] for mg in model_MG_base]
    labels_base = [f'MG {mg} (Base)' for mg in model_MG_base]

    total_trading_curtailment = np.zeros(24)
    for mg in data_to_plot_trading:
        total_trading_curtailment += data_to_plot_trading[mg]

    # ==========================================
    # PART 4: PLOTTING
    # ==========================================
    plt.figure(figsize=(12, 6))

    plt.stackplot(hours, y_values_base, labels=labels_base, alpha=0.6)
    plt.plot(hours, total_trading_curtailment, color='black', linestyle='-', linewidth=2.5, 
             marker='o', markersize=4, label='Total Curtailment (Trading)')

    plt.title('Comparison of RE Curtailment: Base Case vs. P2P Trading', fontsize=14, pad=15)
    plt.xlabel('Hour of the Day', fontsize=12)
    plt.ylabel('Curtailed Power (MW)', fontsize=12)
    plt.xticks(hours)
    plt.xlim(0, 23)  
    plt.grid(True, linestyle=':', alpha=0.6)
    plt.legend(loc='upper right')

    plt.tight_layout()
    plt.show()

    # ==========================================
    # PART 5: PRINT DETAILED COMPARISON RESULTS
    # ==========================================
    print("\n" + "="*85)
    print(f"{'DETAILED COMPARISON: RE CURTAILMENT & UTILIZATION (MW)':^85}")
    print("="*85)
    
    for mg in model_MG:
        print(f"\n[{'='*10} MICROGRID {mg} {'='*10}]")
        
        for time in model_MG[mg].time:
            has_printed_time = False # Cờ để chỉ in mốc thời gian 1 lần nếu có dữ liệu
            
            # Kiểm tra Node PV
            for node in model_MG[mg].node_has_PV:
                peak = model_MG[mg].P_PV_peak[node, time]
                act_base = model_MG_base[mg].P_PV[node, time].value
                act_trad = model_MG[mg].P_PV[node, time].value
                
                curt_base = max(0, peak - act_base) if (peak - act_base) > 0.0001 else 0
                curt_trad = max(0, peak - act_trad) if (peak - act_trad) > 0.0001 else 0
                re_utilized = curt_base - curt_trad # Lượng năng lượng được "cứu" nhờ Trading
                
                # CHỈ IN NẾU BASE CÓ CẮT GIẢM (> 0)
                if curt_base > 0:
                    if not has_printed_time:
                        print(f"\n--- Time {time}:")
                        has_printed_time = True
                    print(f"  Node {node} (PV)   | Base Curtail: {curt_base:.4f} | Trading Curtail: {curt_trad:.4f} | RE Utilized: +{re_utilized:.4f}")
            
            # Kiểm tra Node Wind
            for node in model_MG[mg].node_has_Wind:
                peak = model_MG[mg].P_Wind_peak[node, time]
                act_base = model_MG_base[mg].P_Wind[node, time].value
                act_trad = model_MG[mg].P_Wind[node, time].value
                
                curt_base = max(0, peak - act_base) if (peak - act_base) > 0.0001 else 0
                curt_trad = max(0, peak - act_trad) if (peak - act_trad) > 0.0001 else 0
                re_utilized = curt_base - curt_trad
                
                # CHỈ IN NẾU BASE CÓ CẮT GIẢM (> 0)
                if curt_base > 0:
                    if not has_printed_time:
                        print(f"\n--- Time {time}:")
                        has_printed_time = True
                    print(f"  Node {node} (Wind) | Base Curtail: {curt_base:.4f} | Trading Curtail: {curt_trad:.4f} | RE Utilized: +{re_utilized:.4f}")
def plot_loss_and_load_comparison(model_base, model_trade, trading_line):
    for MG in model_trade.keys():
        print(f"\n--- Đang xử lý và vẽ biểu đồ cho MG {MG} ---")
        
        # 1. Lấy dữ liệu load (MW) và loss (kW)
        load_base, loss_base = print_total_loss_data(model_base[MG], trading_line, MG)
        load_trade, loss_trade = print_total_loss_data(model_trade[MG], trading_line, MG)
        
        # 2. TÍNH TOÁN % TỔN THẤT SO VỚI TẢI
        # Công thức: (Loss_kW / (Load_MW * 1000)) * 100
        # Lưu ý: Kiểm tra load > 0 để tránh lỗi chia cho 0
        loss_pct_base = [(ls / (ld)) * 100 if ld != 0 else 0 for ls, ld in zip(loss_base, load_base)]
        loss_pct_trade = [(ls / (ld)) * 100 if ld != 0 else 0 for ls, ld in zip(loss_trade, load_trade)]
        
        # 3. Khởi tạo trục x (24 giờ)
        hours = list(range(24))
        
        # 4. Tạo Figure và Trục trái (ax1) cho % Loss
        fig, ax1 = plt.subplots(figsize=(10, 6))
        
        # Vẽ % Loss lên trục trái (ax1)
        line1, = ax1.plot(hours, loss_pct_base, label='% Loss Base', marker='', linestyle='-', color='#d62728', linewidth=2)
        line2, = ax1.plot(hours, loss_pct_trade, label='% Loss Trade', marker='', linestyle='--', color="#c94444", linewidth=2)
        
        # Thiết lập trục trái (ax1)
        ax1.set_xlabel("Time step (h)", fontsize=12)
        ax1.set_ylabel("Percentage of loss (%)", fontsize=12, color='#d62728') # Đổi đơn vị thành %
        ax1.set_xticks(hours)
        ax1.grid(True, linestyle='--', alpha=0.4)
        
        # 5. Tạo Trục phải (ax2) cho Load (MW)
        ax2 = ax1.twinx()
        line3, = ax2.plot(hours, load_base, label='Load', marker='s', linestyle='-', color='#1f77b4', linewidth=2, alpha=0.7)
        
        # Thiết lập trục phải (ax2)
        ax2.set_ylabel("Load (MW)", fontsize=12, color='#1f77b4')
        
        # 6. Gộp chung Legend
        lines = [line1, line2, line3]
        labels = [l.get_label() for l in lines]
        ax1.legend(lines, labels, loc='upper center', bbox_to_anchor=(0.5, -0.12), ncol=3, fontsize=11)
        
        # 7. Căn chỉnh và hiển thị
        # plt.title(f"So sánh Load (MW) và Tỉ lệ % Loss trong 24h - Microgrid {MG}", fontsize=14, fontweight='bold', pad=15)
        
        fig.tight_layout() 
        plt.show()
    return
def generate_loss_summary_table(model_base, model_trade, trading_line):
    summary_data = []
    
    # Khởi tạo biến tích lũy cho dòng TOTAL
    t_load_base = 0
    t_loss_base = 0
    t_load_trade = 0
    t_loss_trade = 0

    for MG in model_base.keys():
        # 1. Lấy dữ liệu Load (MW) và Loss (kW)
        load_b, loss_b = print_total_loss_data(model_base[MG], trading_line, MG)
        load_t, loss_t = print_total_loss_data(model_trade[MG], trading_line, MG)
        
        # Tổng cộng trong ngày cho MG này
        sum_ld_b, sum_ls_b = sum(load_b), sum(loss_b)
        sum_ld_t, sum_ls_t = sum(load_t), sum(loss_t)
        
        # 2. Tính % Loss so với tải (Công thức: Loss_kW / (Load_MW * 1000) * 100)
        pct_ls_ld_base = (sum_ls_b / (sum_ld_b) * 100) if sum_ld_b != 0 else 0
        pct_ls_ld_trade = (sum_ls_t / (sum_ld_t) * 100) if sum_ld_t != 0 else 0
        
        # 3. Tính độ lệch tuyệt đối và % cải thiện của riêng lượng Loss
        reduction_kw = pct_ls_ld_base - pct_ls_ld_trade
        pct_improvement = (reduction_kw / sum_ls_b * 100) if sum_ls_b != 0 else 0
        
        # Tích lũy cho TOTAL
        t_load_base += sum_ld_b
        t_loss_base += sum_ls_b
        t_load_trade += sum_ld_t
        t_loss_trade += sum_ls_t
        
        summary_data.append({
            "MG": MG,
            "Base Loss (MWh)": round(sum_ls_b, 2),
            "Trade Loss (MWh)": round(sum_ls_t, 2),
            "Load (MWh)": round(sum_ld_b,2),
            "% Loss/Load (Base)": f"{pct_ls_ld_base:.3f}%",
            "% Loss/Load (Trade)": f"{pct_ls_ld_trade:.3f}%",
            "% Change": round(reduction_kw, 2),
        })

    # 4. Tính toán dòng TOTAL
    total_pct_b = (t_loss_base / (t_load_base) * 100) if t_load_base != 0 else 0
    total_pct_t = (t_loss_trade / (t_load_trade) * 100) if t_load_trade != 0 else 0
    total_reduction = total_pct_b -  total_pct_t
    total_load = t_load_base
    # total_pct_change = (total_reduction / t_loss_base * 100) if t_loss_base != 0 else 0
    
    summary_data.append({
        "MG": "TOTAL",
        "Base Loss (MWh)": round(t_loss_base, 2),
        "Trade Loss (MWh)": round(t_loss_trade, 2),
        "Load (MWh)": round(total_load,2),
        "% Loss/Load (Base)": f"{total_pct_b:.3f}%",
        "% Loss/Load (Trade)": f"{total_pct_t:.3f}%",
        "% Change": round(total_reduction, 2),
    })

    # 5. Hiển thị bảng
    df = pd.DataFrame(summary_data)
    print("\n" + "="*95)
    print("BẢNG PHÂN TÍCH TỔN THẤT VÀ TỈ LỆ % TRÊN TẢI")
    print("="*95)
    print(df.to_string(index=False))
    print("="*95)
    
    return df
def plot_history_r_combined(history_r):
    pairs = sorted(list(set([(k[0], k[1]) for k in history_r.keys() if k[0] < k[1]])))
    
    if not pairs:
        print("Không có dữ liệu thỏa mãn điều kiện i < j.")
        return

    num_pairs = len(pairs)
    fig, axes = plt.subplots(num_pairs, 1, figsize=(13, 7 * num_pairs), squeeze=False)
    
    cmap = plt.get_cmap('jet') 
    colors = [cmap(i) for i in np.linspace(0, 1, 24)]
    
    # Ngưỡng tối thiểu để hiển thị trên thang log (khớp với ylim bottom)
    BOTTOM_THRESHOLD = 1e-9

    for idx, (i, j) in enumerate(pairs):
        ax = axes[idx, 0]
        
        # --- PHẦN 1: TÍNH TOÁN THỐNG KÊ ---
        all_steps = sorted(list(set([k[3] for k in history_r.keys() if k[0] == i and k[1] == j])))
        
        mean_vals, min_vals, max_vals = [], [], []
        
        for s in all_steps:
            vals_at_step = [v for k, v in history_r.items() if k[0] == i and k[1] == j and k[3] == s]
            if vals_at_step:
                # Ép giá trị không thấp hơn ngưỡng để luôn thấy đường trên đồ thị
                mean_vals.append(max(np.mean(vals_at_step), BOTTOM_THRESHOLD))
                min_vals.append(max(np.min(vals_at_step), BOTTOM_THRESHOLD))
                max_vals.append(max(np.max(vals_at_step), BOTTOM_THRESHOLD))
            else:
                mean_vals.append(np.nan)
                min_vals.append(np.nan)
                max_vals.append(np.nan)

        # --- PHẦN 2: VẼ CÁC ĐƯỜNG CÁ THỂ (MỜ) ---
        available_t = sorted(list(set([k[2] for k in history_r.keys() if k[0] == i and k[1] == j])))
        for t in available_t:
            data_t = sorted([(k[3], v) for k, v in history_r.items() if k[0] == i and k[1] == j and k[2] == t])
            if data_t:
                steps_t, residuals_t = zip(*data_t)
                # Dùng clip để các đường nhỏ không bị mất hút dưới đáy
                residuals_t_clipped = np.maximum(residuals_t, BOTTOM_THRESHOLD)
                ax.plot(steps_t, residuals_t_clipped, color=colors[int(t % 24)], alpha=0.15, linewidth=0.6)

        # --- PHẦN 3: VẼ VÙNG XU HƯỚNG & ĐƯỜNG MEAN NỔI BẬT ---
        # Vẽ vùng Min-Max mờ phía sau
        ax.fill_between(all_steps, min_vals, max_vals, color='gray', alpha=0.1, label='Min-Max Range')
        
        # ĐƯỜNG MEAN: Hiệu ứng Glow (tỏa sáng)
        # 1. Lớp viền trắng ngoài cùng để tách biệt với các đường cá thể
        ax.plot(all_steps, mean_vals, color='white', linewidth=6, zorder=10) 
        # 2. Đường chính màu đỏ đậm/cam để cực kỳ nổi bật
        ax.plot(all_steps, mean_vals, color='#FF3300', linewidth=3, 
                linestyle='-', marker='o', markevery=max(1, len(all_steps)//20), 
                markersize=6, label='Mean Residual (Trend)', zorder=11)

        # --- CẤU HÌNH ĐỒ THỊ ---
        ax.set_title(f'Residual Convergence Trend: Pair (MG_{i}, MG_{j})', fontsize=15, fontweight='bold', pad=15)
        ax.set_xlabel('Step (Iteration)', fontsize=12)
        ax.set_ylabel('Residual (pu)', fontsize=12)
        ax.set_yscale('log')
        # ax.set_yscale('linear')
        
        # Fix cứng ylim để đường chạm đáy vẫn nằm trên trục hoành
        ax.set_ylim(bottom=BOTTOM_THRESHOLD, top=max(max_vals)*2 if max_vals else 1) 
        
        ax.grid(True, which="both", ls="--", alpha=0.3)
        
        # Đưa Legend ra ngoài hoặc góc rõ ràng
        ax.legend(loc='upper right', frameon=True, shadow=True, fontsize='medium')

    plt.tight_layout()
    plt.show()
    return
def print_obj_cost(model_base, model_trade):
    total_base = 0
    total_trade = 0
    
    # Tiêu đề bảng
    header = f"{'Microgrid':<15} | {'Base Cost':>12} | {'Trade Cost':>12} | {'Saving':>12} | {'% Change':>10}"
    print("-" * len(header))
    print(header)
    print("-" * len(header))

    for MG in model_base.keys():
        cost_base = value(model_base[MG].OBJ.expr)
        cost_trade = value(model_trade[MG].OBJ.expr)
        
        total_base += cost_base
        total_trade += cost_trade
        
        diff = cost_base - cost_trade
        # Tránh chia cho 0 nếu cost_base = 0
        p_change = (diff / cost_base * 100) if cost_base != 0 else 0
        
        # In từng dòng với định dạng số thập phân 2 chữ số
        print(f"{MG:<15} | {cost_base:>12.2f} | {cost_trade:>12.2f} | {diff:>12.2f} | {p_change:>9.2f}%")

    # In dòng tổng kết
    total_diff = total_base - total_trade
    total_p_change = (total_diff / total_base * 100) if total_base != 0 else 0
    
    print("-" * len(header))
    print(f"{'TOTAL':<15} | {total_base:>12.2f} | {total_trade:>12.2f} | {total_diff:>12.2f} | {total_p_change:>9.2f}%")
    print("-" * len(header))
    
    # Nhận xét nhanh
    if total_diff > 0:
        print(f"==> Kết quả: Trading giúp giảm được {total_diff:.2f} đơn vị chi phí ({total_p_change:.2f}%).")
    else:
        print(f"==> Kết quả: Trading không làm giảm tổng chi phí.")
    return 
def show_voltage_profiles(models_dict, case_name):
    """
    Plot and display individual voltage profiles for each microgrid.
    
    Parameters:
    models_dict (dict): Dictionary containing models {mg_id: model_obj}
    case_name (str): Scenario name (e.g., 'Base' or 'Trading')
    """
    for mg_id, mg_model in models_dict.items():
        # 1. Extract data from the model
        nodes = list(mg_model.index_of_node)
        times = list(mg_model.time)
        
        data_list = []
        for t in times:
            row = []
            for n in nodes:
                val = mg_model.U_node[n, t].value
                # Calculate square root to convert to p.u. if the model uses U^2
                u_val = np.sqrt(val) if (val is not None and val > 0) else 1.0
                row.append(u_val)
            data_list.append(row)
        
        df_voltage = pd.DataFrame(data_list, columns=nodes)
        
        # 2. Compute statistical profiles
        v_mean = df_voltage.mean(axis=0)
        v_min = df_voltage.min(axis=0)
        v_max = df_voltage.max(axis=0)
        node_indices = np.arange(1, len(nodes) + 1)

        # 3. Generate independent figure
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Plot voltage fluctuation envelope
        ax.fill_between(node_indices, v_min, v_max, color='lightgreen', alpha=0.3, label='Voltage Swing')
        
        # Plot mean voltage trajectory
        ax.plot(node_indices, v_mean, color='blue', lw=1.5, label='Mean', marker='s', markersize=3)
        
        # Plot operational limits
        ax.axhline(y=1.05, color='crimson', linestyle='--', alpha=0.7, label='Limit 1.05')
        ax.axhline(y=0.95, color='crimson', linestyle='--', alpha=0.7, label='Limit 0.95')
        ax.axhline(y=1.0, color='black', lw=0.8, alpha=0.5)

        # Figure formatting
        # ax.set_title(f'Microgrid {mg_id} - {case_name} Scenario', fontsize=14, fontweight='bold')
        ax.set_xlabel('Node Index', fontsize=12)
        ax.set_ylabel('Voltage (pu)', fontsize=12)
        ax.set_ylim(0.94, 1.06)
        ax.grid(True, linestyle=':', alpha=0.6)
        ax.legend(loc='lower left', fontsize='medium')

        plt.tight_layout()
        plt.show()
def BESS_SOC_plot(model):
    # Khởi tạo khung vẽ
    plt.figure(figsize=(10, 6))
    
    # Duyệt qua các Microgrid (mg) trong model
    for mg in model:
        # Duyệt qua từng nút có lắp đặt BESS
        for node in model[mg].node_has_BESS:
            # Lấy danh sách thời gian và sắp xếp theo thứ tự tăng dần
            times = sorted(list(model[mg].time))
            
            # Trích xuất giá trị SOC tương ứng với từng mốc thời gian
            # value(...) dùng để lấy giá trị số từ biến của Pyomo
            soc_values = [value(model[mg].SOC[node, t]) for t in times]
            
            # Vẽ đường SOC cho nút hiện tại
            # label sẽ giúp phân biệt các đường trong chú thích (legend)
            plt.plot(times, soc_values, marker='o', linestyle='-', label=f"MG:{mg}")

    # Cấu hình các thông tin bổ trợ cho đồ thị
    plt.xlabel('Time Step (h)', fontsize=12)
    plt.ylabel('SOC', fontsize=12)
    # plt.title('Biểu đồ State of Charge (SOC) của hệ thống BESS', fontsize=14)
    
    # Hiển thị chú thích để phân biệt 3 đường
    plt.legend()
    
    # Thêm lưới để dễ quan sát
    plt.grid(True, linestyle='--', alpha=0.7)
    
    # Hiển thị hoặc lưu đồ thị
    plt.tight_layout()
    plt.show()
    # plt.savefig('BESS_SOC_Plot.png') # Bỏ comment nếu muốn lưu file
def plot_bess_degradation_stacked(model):
    fix_value = scale_power / S_base
    results_per_mg = {}

    for mg in model:
        if mg == 0: 
           continue
        mg_hourly_degradation = []
        for t in model[mg].time:
            hourly_sum = 0
            for node in model[mg].node_has_BESS:
                # Tính toán chi phí suy hao dựa trên PC_BESS (nạp) và PD_BESS (xả)
                val = model[mg].Theta_BESS * (
                    model[mg].Eff_BESS * model[mg].PC_BESS[node, t].value / fix_value + 
                    model[mg].PD_BESS[node, t].value / (fix_value * model[mg].Eff_BESS)
                )
                hourly_sum += val
            mg_hourly_degradation.append(hourly_sum)
        results_per_mg[mg] = np.array(mg_hourly_degradation) # Chuyển sang numpy array để cộng bottom cho dễ

    # --- 2. CHUẨN BỊ THÔNG SỐ VẼ ---
    first_mg = list(model.keys())[0]
    hours = np.array(list(model[first_mg].time))
    labels = [f"MG {mg}" for mg in results_per_mg.keys()]

    # Bảng màu Tableau 10 (giống hệt trong hình mẫu)
    colors = ['#4e79a7', '#f28e2c', '#59a14f', '#e15759', '#76b7b2', '#54a24b', '#edc948']

    # --- 3. VẼ BIỂU ĐỒ CỘT CHỒNG ---
    fig, ax = plt.subplots(figsize=(12, 6), dpi=100)

    # Khởi tạo mảng bottom bằng 0 để bắt đầu xếp chồng
    bottom = np.zeros(len(hours))
    bar_width = 0.7  # Độ rộng của cột (0.7-0.8 nhìn sẽ thoáng)

    for i, mg in enumerate(results_per_mg):
        values = results_per_mg[mg]
        ax.bar(hours, values, 
              bottom=bottom, 
              label=labels[i], 
              color=colors[i % len(colors)],
              width=bar_width,
              edgecolor='white', # Tạo viền trắng mảnh giữa các block cho sạch
              linewidth=0.5,
              alpha=0.85)
        
        # Cập nhật bottom cho lớp tiếp theo
        bottom += values

    # --- 4. ĐỊNH DẠNG ---
    ax.set_title("BESS Degradation Cost: Multi-Microgrid System", 
                fontsize=14, fontweight='bold', pad=20)
    ax.set_xlabel("Hour of the Day", fontsize=11)
    ax.set_ylabel("Degradation Cost (cents)", fontsize=11)

    # Định dạng trục X và Y
    ax.set_xticks(hours)
    ax.set_xlim(-0.5, 23.5) # Căn chỉnh để cột đầu và cuối không bị dính lề
    ax.set_ylim(0, 22)

    # Lưới chấm mờ
    ax.grid(axis='y', linestyle=':', color='gray', alpha=0.3)

    # Chú giải
    ax.legend(loc='upper right', frameon=True, facecolor='white', framealpha=0.9)

    plt.tight_layout()
    plt.show()
    return results_per_mg
def plot_price(model, lamda, trading_line):
    # Khai báo trực tiếp dữ liệu giá lưới bên trong hàm
    price_buy = [15.59, 15.08, 15.21, 15.08, 14.69, 15.59, 16.23, 16.87, 17.26, 17.51, 
                 17.77, 18.03, 18.28, 18.03, 18.03, 17.9, 17.9, 17.77, 17.64, 17.51, 
                 17.13, 15.85, 15.59, 15.21]

    price_sell = [13.59, 13.08, 13.21, 13.08, 12.69, 13.59, 14.23, 14.87, 15.26, 15.51, 
                  15.77, 16.03, 16.28, 16.03, 16.03, 15.9, 15.9, 15.77, 15.64, 15.51, 
                  15.13, 13.85, 13.59, 13.21]

    plt.figure(figsize=(12, 6))
    
    # 1. Vẽ các đường giá giao dịch giữa các Microgrids (P2P)
    for MG in model:
        i = MG
        for node in model[MG].node_has_trade:
            for j in trading_line[i, node]:
                if (i > j): # Quy định i > j để tránh vẽ lặp đường giao dịch hai chiều
                    continue
                
                time_axis = []
                price_axis = []
                
                for t in model[MG].time:
                    price_temp = lamda[(i, j, t)]
                    time_axis.append(t)
                    price_axis.append(price_temp)
                
                plt.plot(time_axis, price_axis, marker='o', linestyle='-', label=f'MG{i} ➔ MG{j}')

    # 2. Lấy trục thời gian từ model để vẽ giá lưới (Utility Grid)
    # Giả định tất cả các MG có cùng trục thời gian t
    sample_mg = next(iter(model))
    time_steps = list(model[sample_mg].time)
    
    # Vẽ đường Price Buy (Giá mua từ lưới)
    plt.plot(time_steps, price_buy, color='red', linestyle='--', linewidth=3, 
             label='Buy price from grid ($P_{buy}$)')
    
    # Vẽ đường Price Sell (Giá bán lên lưới)
    plt.plot(time_steps, price_sell, color='red', linestyle='--', linewidth=3, 
             label='Buy price to grid  ($P_{sell}$)')

    # Cấu hình giao diện đồ thị
    # plt.title('Biểu đồ so sánh giá giao dịch P2P và giá lưới', fontsize=14)
    plt.xlabel('Time step (h)', fontsize=12)
    plt.ylabel('Trading price (cents)', fontsize=12)
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left') 
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.tight_layout()
    
    plt.show()
    return 
def check_socp_relaxation_gap(global_model_MG, tolerance=1e-4):
    print("\n" + "="*80)
    print(f"{'BẢNG KIỂM TRA ĐỘ LỆCH SOCP (SOCP RELAXATION GAP)':^80}")
    print("="*80)
    print(f"{'MG':<5} | {'Time':<5} | {'Line (i -> j)':<15} | {'I_line (l)':<12} | {'(P^2+Q^2)/U':<12} | {'Gap Error':<12}")
    print("-" * 80)
    
    max_gap_overall = 0
    total_violations = 0

    for mg_id, model in global_model_MG.items():
        for t in model.time:
            for i in model.index_of_node:
                # Duyệt qua các nút con (j) của nút cha (i)
                for j in model.child_of[i]:
                    if i != j: # Đảm bảo là nhánh nối giữa 2 nút khác nhau
                        # Lấy giá trị biến từ model sau khi giải
                        try:
                            p_val = pe.value(model.P_line[j, t])
                            q_val = pe.value(model.Q_line[j, t])
                            i_val = pe.value(model.I_line[j, t])   # Đây chính là biến l (bình phương dòng điện)
                            u_val = 1  # U_node của nút cha (V^2)
                        except ValueError:
                            # Nếu biến chưa được gán giá trị (chưa solve hoặc lỗi)
                            continue
                        
                        # Kiểm tra U_node để tránh chia cho 0
                        if u_val is not None and u_val > 0.01:
                            right_side = (p_val**2 + q_val**2) / u_val
                            gap = i_val - right_side
                            
                            if abs(gap) > max_gap_overall:
                                max_gap_overall = abs(gap)
                            if abs(gap) > tolerance:
                                total_violations += 1
                                print(f"{mg_id:<5} | {t:<5} | {i} -> {j:<10} | {i_val:<12.5f} | {right_side:<12.5f} | {gap:<12.5e}")
                                
    print("-" * 80)
    print(f"Tổng số nhánh có Gap vượt ngưỡng {tolerance}: {total_violations}")
    print(f"Độ lệch lớn nhất toàn hệ thống (Max Gap): {max_gap_overall:.5e}")
    print("="*80)
    return max_gap_overall
#! GLOBAL DATA
def create_trading_information(filename):
    df =pd.read_excel(filename, engine= "openpyxl")
    trading_list = {}
    trading_node = {}
    for index, value in df['MG_from'].items():
        MG_from	= value
        if pd.isna(value) or pd.isna(df.at[index, 'node_from']) or pd.isna(df.at[index, 'MG_to']) or pd.isna(df.at[index, 'node_to']):
            break
        node_from =int(df.at[index, 'node_from'])	
        MG_to	=int(df.at[index, 'MG_to'])
        node_to =int(df.at[index, 'node_to'])

        trading_list.setdefault((MG_from, node_from), []).append((MG_to))
        trading_node.setdefault((MG_from, node_from), []).append((node_to))
        
    return trading_list,trading_node
def load_trading_network(number_of_MG):
    trading_line = {}
    trading_node = {}
    for MG in number_of_MG:
        filename = f"./Trading_data/MG{MG}.xlsx"
        lines, nodes = create_trading_information(filename)
        for key, val in lines.items():
            trading_line.setdefault(key, []).extend(val)
        for key, val in nodes.items():
            trading_node.setdefault(key, []).extend(val)
    return trading_line, trading_node
def initialize_admm_params(model_MG, trading_line):
    """Khởi tạo p_trade, lamda và các biến trạng thái ADMM."""
    p_trade = {}
    lamda_ATC = {}
    rho_ATC = {}
    P_shed_critical = {}
    P_shed_noncritical = {}
    lamda = {}
    kiem_tra_dung = {}
    for i, h in trading_line.keys():
        for j in trading_line[i, h]:
            for t in range(24):
                model_MG[i].phat_ADMM[h, t] = 0.05
                lamda[(i, j, t)] = model_MG[i].Price_sell[t].value
                kiem_tra_dung[(i, j, t)] = 0
    return lamda_ATC, rho_ATC, P_shed_critical, P_shed_noncritical, lamda, kiem_tra_dung
def push_and_solve_admm(model, mg_id, lamda, p_trade_external, trading_line, epsilon=1e-6):
    """
    Cập nhật dữ liệu từ các MG lân cận, giải mô hình và trả về kết quả P_trade.
    """
    # 1. Cập nhật tham số (Gán giá trị cho Param hoặc Var cố định)
    for node in model.node_has_trade:
        for t in model.time:
            neighbors = trading_line.get((mg_id, node), [])
            for j in neighbors:
                model.trading_cost[node, t] = lamda[(mg_id, j, t)]
                val_p = p_trade_external[(j, mg_id, t)]
                model.P_form_another_MG[node, t] = val_p[0] if isinstance(val_p, list) else val_p

    # 2. Giải mô hình
    results = solve(model)
    # 3. Kiểm tra hội tụ và làm sạch dữ liệu (Chỉ quét các biến tích cực)
    if results.solver.termination_condition == pe.TerminationCondition.optimal:
        # Chỉ làm sạch các biến Var có giá trị gần bằng 0
        for v in model.component_data_objects(pe.Var, active=True):
            if v.value is not None and abs(v.value) < epsilon:
                v.set_value(0)
    else:
        print(f"⚠️ CẢNH BÁO: MG {mg_id} dừng với trạng thái: {results.solver.termination_condition}")

    # 4. Trích xuất kết quả bằng vòng lặp for truyền thống
    results_dict = {}  # Khởi tạo một dictionary trống
    
    for node in model.node_has_trade:
        for t in model.time:
            neighbors = trading_line.get((mg_id, node), [])
            for j in neighbors:
                p_val = 0 # Đã xóa P_trade
                key = (mg_id, j, t)
                results_dict[key] = p_val
    return results_dict

#! MÔ PHỔNG SỰ CỐ
def apply_custom_fault(model, export_limit, config):
    model.fault_constraints = pe.ConstraintList()
    if not config:
        return
        
    if isinstance(config, dict):
        config_list = [{
            'start_time': config.get('time_fault', 17),
            'end_time': config.get('end_time', 99),
            'type': config.get('type', 'grid_loss'),
            'severity': config.get('severity', 0.0)
        }]
    else:
        config_list = config

    t_start = getattr(model, 't_start', 0)
    for fault in config_list:
        start_time = fault.get('start_time', 17)
        end_time = fault.get('end_time', 99)
        fault_type = fault.get('type', 'grid_loss')
        severity = fault.get('severity', 0.0)

        for t_idx in model.time:
            t_real = t_start + t_idx
            if start_time <= t_real <= end_time:
                if fault_type == 'grid_loss':
                    model.fault_constraints.add(model.P_line[0, t_idx] == 0)
                    model.fault_constraints.add(model.Q_line[0, t_idx] == 0)
                elif fault_type == 'pv_loss':
                    for node in model.node_has_PV:
                        model.fault_constraints.add(model.P_PV[node, t_idx] <= severity * model.P_PV_peak[node, t_idx])
                elif fault_type == 'der_loss':
                    if hasattr(model, 'node_has_PV'):
                        for node in model.node_has_PV:
                            model.fault_constraints.add(model.P_PV[node, t_idx] <= severity * model.P_PV_peak[node, t_idx])
                    if hasattr(model, 'node_has_Wind'):
                        for node in model.node_has_Wind:
                            model.fault_constraints.add(model.P_Wind[node, t_idx] <= severity * model.P_Wind_peak[node, t_idx])
    return
